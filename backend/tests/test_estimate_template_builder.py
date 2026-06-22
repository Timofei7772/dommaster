from openpyxl import load_workbook


def test_build_template_creates_required_sheets(tmp_path):
    from app.services.estimate_template_builder import build_estimate_template

    output = tmp_path / "estimate_template.xlsx"
    build_estimate_template(output)

    assert output.exists()


def test_classify_section_uses_keyword_rules():
    from app.services.estimate_template_builder import classify_section

    rules = [
        {
            "priority": 10,
            "keywords_csv": "плитка,сантехника",
            "section_name": "Ванная комната",
            "is_active": True,
        },
    ]

    assert classify_section("Монтаж плитки и сантехники", rules) == "Ванная комната"


def test_vat_mode_flag_switches_formula_strategy(tmp_path):
    from app.services.estimate_template_builder import build_estimate_template

    output = tmp_path / "estimate_template.xlsx"
    build_estimate_template(output)
    wb = load_workbook(output, data_only=False)

    settings = wb["Настройки"]
    smeta = wb["Смета"]

    assert settings["B1"].value is True
    assert "IF(" in str(smeta["F116"].value)


def test_template_contains_visible_and_hidden_sheets(tmp_path):
    from app.services.estimate_template_builder import build_estimate_template

    output = tmp_path / "estimate_template.xlsx"
    build_estimate_template(output)
    wb = load_workbook(output)

    assert wb.sheetnames == ["Смета", "Справочники", "Настройки"]
    assert wb["Справочники"].sheet_state == "hidden"
    assert wb["Настройки"].sheet_state == "hidden"



def test_header_and_table_labels_match_design(tmp_path):
    from app.services.estimate_template_builder import build_estimate_template

    output = tmp_path / "estimate_template.xlsx"
    build_estimate_template(output)
    ws = load_workbook(output)["Смета"]

    assert ws["B2"].value == 'ООО "Название компании"'
    assert ws["C3"].value == "ЛОКАЛЬНАЯ СМЕТА"
    assert ws["A12"].value in (None, "")
    assert ws["F12"].value == "Стоимость"


def test_item_and_section_formula_cells_are_written(tmp_path):
    from app.services.estimate_template_builder import build_estimate_template

    output = tmp_path / "estimate_template.xlsx"
    build_estimate_template(output)
    ws = load_workbook(output, data_only=False)["Смета"]

    assert str(ws["F13"].value).startswith("=IF(")
    assert "SUMIFS" in str(ws["F30"].value)
    assert "Настройки!$B$1" in str(ws["F116"].value)



def test_hidden_helper_columns_exist(tmp_path):
    from app.services.estimate_template_builder import build_estimate_template

    output = tmp_path / "estimate_template.xlsx"
    build_estimate_template(output)
    ws = load_workbook(output)["Смета"]

    assert ws.column_dimensions["G"].hidden is True
    assert ws["G13"].value == "item"



def test_fill_template_groups_rows_by_classified_section(tmp_path):
    from app.services.estimate_template_builder import build_estimate_template, fill_estimate_template

    output = tmp_path / "estimate_template.xlsx"
    rows = [
        {"name": "Демонтаж плитки", "unit": "м2", "quantity": 10, "unit_price": 300},
        {"name": "Монтаж сантехники", "unit": "шт", "quantity": 1, "unit_price": 12000},
    ]

    build_estimate_template(output)
    fill_estimate_template(output, rows)
    ws = load_workbook(output, data_only=False)["Смета"]

    assert ws["B13"].value == "Подготовительные работы"
    assert ws["B17"].value == "Ванная комната"



def test_cli_generates_template_file(tmp_path):
    from scripts.generate_template import main

    output = tmp_path / "estimate_template.xlsx"
    exit_code = main(["--output", str(output)])

    assert exit_code == 0
    assert output.exists()
