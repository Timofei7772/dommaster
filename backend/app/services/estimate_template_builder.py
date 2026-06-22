from __future__ import annotations

import csv
import json
from pathlib import Path
from typing import Any, Iterable

from openpyxl import Workbook, load_workbook


DEFAULT_RULES = [
    {
        "priority": 10,
        "keywords_csv": "демонтаж,вывоз",
        "section_name": "Подготовительные работы",
        "section_code": "prep",
        "is_active": True,
    },
    {
        "priority": 20,
        "keywords_csv": "плитка,сантехника",
        "section_name": "Ванная комната",
        "section_code": "bathroom",
        "is_active": True,
    },
    {
        "priority": 30,
        "keywords_csv": "ламинат,плинтус",
        "section_name": "Жилые комнаты (Пол)",
        "section_code": "floor",
        "is_active": True,
    },
    {
        "priority": 40,
        "keywords_csv": "шпаклевка,обои",
        "section_name": "Стены и потолки",
        "section_code": "walls",
        "is_active": True,
    },
]


ITEM_START_ROW = 13
ITEM_END_ROW = 112
SECTION_TOTAL_ROW = 30
VISIBLE_ROW_LIMIT = 500


def _normalize_text(value: str) -> str:
    return " ".join((value or "").lower().replace("ё", "е").split())


def _keyword_matches(normalized_name: str, keyword: str) -> bool:
    normalized_keyword = _normalize_text(keyword)
    if not normalized_keyword:
        return False
    if normalized_keyword in normalized_name:
        return True

    stem = normalized_keyword.rstrip("аеиоуыэюяйь")
    if len(stem) < 4:
        stem = normalized_keyword

    for token in normalized_name.split():
        if token.startswith(stem):
            return True

    return False


def classify_section(name: str, rules: Iterable[dict] | None = None) -> str:
    normalized_name = _normalize_text(name)
    active_rules = [rule for rule in (rules or DEFAULT_RULES) if rule.get("is_active", True)]

    for rule in sorted(active_rules, key=lambda item: item.get("priority", 0)):
        for keyword in str(rule.get("keywords_csv", "")).split(","):
            if _keyword_matches(normalized_name, keyword):
                return str(rule.get("section_name") or "Прочие работы")

    return "Прочие работы"


def build_estimate_template(output_path: Path | str) -> Path:
    target_path = Path(output_path)
    target_path.parent.mkdir(parents=True, exist_ok=True)

    workbook = Workbook()
    smeta_sheet = workbook.active
    smeta_sheet.title = "Смета"
    references_sheet = workbook.create_sheet("Справочники")
    settings_sheet = workbook.create_sheet("Настройки")

    references_sheet.sheet_state = "hidden"
    settings_sheet.sheet_state = "hidden"

    _populate_settings_sheet(settings_sheet)
    _populate_references_sheet(references_sheet)
    _populate_smeta_sheet(smeta_sheet)

    workbook.save(target_path)
    return target_path


def fill_estimate_template(workbook_path: Path | str, rows: Iterable[dict], header_data: dict | None = None) -> Path:
    target_path = Path(workbook_path)
    workbook = load_workbook(target_path)
    sheet = workbook["Смета"]

    _clear_fill_area(sheet)

    grouped_rows: dict[str, list[dict[str, Any]]] = {}
    for source_row in rows:
        row = dict(source_row)
        section_name = row.get("section_name") or classify_section(str(row.get("name") or ""))
        grouped_rows.setdefault(section_name, []).append(row)

    current_row = ITEM_START_ROW
    for section_index, (section_name, section_rows) in enumerate(grouped_rows.items(), start=1):
        _write_section_row(sheet, current_row, section_index, section_name)
        current_row += 1

        for item_index, row in enumerate(section_rows, start=1):
            _write_item_row(sheet, current_row, section_index, item_index, section_name, row)
            current_row += 1

        _write_section_total_row(sheet, current_row, section_name)
        current_row += 1

        if section_index < len(grouped_rows):
            _write_blank_row(sheet, current_row)
            current_row += 1

    if header_data:
        if header_data.get("company_name"):
            sheet["B2"] = header_data["company_name"]
        if header_data.get("document_title"):
            sheet["C3"] = header_data["document_title"]

    workbook.save(target_path)
    return target_path


def load_rows_from_json(path: Path | str) -> list[dict[str, Any]]:
    with Path(path).open("r", encoding="utf-8") as file:
        data = json.load(file)
    return [_normalize_row(row) for row in data]


def load_rows_from_csv(path: Path | str) -> list[dict[str, Any]]:
    with Path(path).open("r", encoding="utf-8-sig", newline="") as file:
        reader = csv.DictReader(file)
        return [_normalize_row(row) for row in reader]


def _normalize_row(row: dict[str, Any]) -> dict[str, Any]:
    normalized = dict(row)
    for key in ("quantity", "unit_price", "materials_price", "labor_price", "machines_price"):
        value = normalized.get(key)
        if value in (None, ""):
            continue
        normalized[key] = float(value)
    return normalized


def _populate_settings_sheet(sheet) -> None:
    settings = {
        "B1": True,
        "B2": 0.20,
        "B3": "руб.",
        "B4": "СМ-",
        "B11": 'ООО "Название компании"',
        "B12": "+7 (000) 000-00-00",
        "B13": "info@example.ru",
        "B14": "www.example.ru",
        "B15": "Адрес офиса",
        "B16": "ИНН / ОГРН",
    }
    for cell, value in settings.items():
        sheet[cell] = value


def _populate_references_sheet(sheet) -> None:
    headers = ["priority", "keywords_csv", "section_name", "section_code", "room_type", "is_active"]
    for index, header in enumerate(headers, start=1):
        sheet.cell(row=1, column=index, value=header)

    for row_index, rule in enumerate(DEFAULT_RULES, start=2):
        sheet.cell(row=row_index, column=1, value=rule["priority"])
        sheet.cell(row=row_index, column=2, value=rule["keywords_csv"])
        sheet.cell(row=row_index, column=3, value=rule["section_name"])
        sheet.cell(row=row_index, column=4, value=rule["section_code"])
        sheet.cell(row=row_index, column=6, value=rule["is_active"])


def _populate_smeta_sheet(sheet) -> None:
    sheet["B2"] = 'ООО "Название компании"'
    sheet["C3"] = "ЛОКАЛЬНАЯ СМЕТА"
    sheet["A12"] = ""
    sheet["B12"] = "Наименование работ"
    sheet["C12"] = "Ед. изм."
    sheet["D12"] = "Кол-во"
    sheet["E12"] = "Цена за ед."
    sheet["F12"] = "Стоимость"

    for column_name in ("G", "H", "I", "J"):
        sheet.column_dimensions[column_name].hidden = True

    for row_index in range(ITEM_START_ROW, ITEM_END_ROW + 1):
        row_kind = "section_total" if row_index == SECTION_TOTAL_ROW else "item"
        section_name = "Подготовительные работы"

        sheet[f"G{row_index}"] = row_kind
        sheet[f"H{row_index}"] = section_name
        sheet[f"I{row_index}"] = "demo"
        sheet[f"J{row_index}"] = f"row-{row_index}"
        sheet[f"F{row_index}"] = (
            _item_formula(row_index)
            if row_kind == "item"
            else _section_total_formula(row_index)
        )

    sheet["E114"] = "Итого по разделам"
    sheet["F114"] = "=ROUND(SUMIFS($F$13:$F$500,$G$13:$G$500,\"section_total\"),2)"
    sheet["E115"] = "Итого без НДС"
    sheet["F115"] = "=IF(Настройки!$B$1,$F$114,ROUND($F$117-$F$116,2))"
    sheet["E116"] = "НДС"
    sheet["F116"] = "=IF(Настройки!$B$1,ROUND($F$115*Настройки!$B$2,2),ROUND($F$117-($F$117/(1+Настройки!$B$2)),2))"
    sheet["E117"] = "Всего по смете"
    sheet["F117"] = "=IF(Настройки!$B$1,ROUND($F$115+$F$116,2),$F$114)"


def _clear_fill_area(sheet) -> None:
    for row_index in range(ITEM_START_ROW, VISIBLE_ROW_LIMIT + 1):
        for column_name in ("A", "B", "C", "D", "E", "F", "G", "H", "I", "J"):
            sheet[f"{column_name}{row_index}"] = None

    sheet["E114"] = "Итого по разделам"
    sheet["F114"] = "=ROUND(SUMIFS($F$13:$F$500,$G$13:$G$500,\"section_total\"),2)"
    sheet["E115"] = "Итого без НДС"
    sheet["F115"] = "=IF(Настройки!$B$1,$F$114,ROUND($F$117-$F$116,2))"
    sheet["E116"] = "НДС"
    sheet["F116"] = "=IF(Настройки!$B$1,ROUND($F$115*Настройки!$B$2,2),ROUND($F$117-($F$117/(1+Настройки!$B$2)),2))"
    sheet["E117"] = "Всего по смете"
    sheet["F117"] = "=IF(Настройки!$B$1,ROUND($F$115+$F$116,2),$F$114)"


def _write_section_row(sheet, row_index: int, section_index: int, section_name: str) -> None:
    sheet[f"A{row_index}"] = section_index
    sheet[f"B{row_index}"] = section_name
    sheet[f"G{row_index}"] = "section"
    sheet[f"H{row_index}"] = section_name
    sheet[f"I{row_index}"] = "section"
    sheet[f"J{row_index}"] = f"section-{section_index}"


def _write_item_row(sheet, row_index: int, section_index: int, item_index: int, section_name: str, row: dict[str, Any]) -> None:
    quantity = float(row.get("quantity") or 1)
    unit_price = _resolve_unit_price(row)

    sheet[f"A{row_index}"] = row.get("item_number") or f"{section_index}.{item_index}"
    sheet[f"B{row_index}"] = row.get("name") or ""
    sheet[f"C{row_index}"] = row.get("unit") or "шт"
    sheet[f"D{row_index}"] = quantity
    sheet[f"E{row_index}"] = unit_price
    sheet[f"F{row_index}"] = _item_formula(row_index)
    sheet[f"G{row_index}"] = "item"
    sheet[f"H{row_index}"] = section_name
    sheet[f"I{row_index}"] = row.get("section_name") or "classified"
    sheet[f"J{row_index}"] = row.get("source_id") or f"item-{section_index}-{item_index}"


def _write_section_total_row(sheet, row_index: int, section_name: str) -> None:
    sheet[f"B{row_index}"] = "Итого по разделу"
    sheet[f"F{row_index}"] = _section_total_formula(row_index)
    sheet[f"G{row_index}"] = "section_total"
    sheet[f"H{row_index}"] = section_name
    sheet[f"I{row_index}"] = "subtotal"
    sheet[f"J{row_index}"] = f"subtotal-{row_index}"


def _write_blank_row(sheet, row_index: int) -> None:
    sheet[f"G{row_index}"] = "blank"
    sheet[f"I{row_index}"] = "spacer"
    sheet[f"J{row_index}"] = f"blank-{row_index}"


def _item_formula(row_index: int) -> str:
    return f'=IF($G{row_index}="item",IF(OR($D{row_index}="",$E{row_index}=""),"",ROUND($D{row_index}*$E{row_index},2)),"")'


def _section_total_formula(row_index: int) -> str:
    return f'=IF($G{row_index}="section_total",ROUND(SUMIFS($F$13:$F$500,$G$13:$G$500,"item",$H$13:$H$500,$H{row_index}),2),"")'


def _resolve_unit_price(row: dict[str, Any]) -> float:
    if row.get("unit_price") not in (None, ""):
        return float(row["unit_price"])

    total = 0.0
    for key in ("materials_price", "labor_price", "machines_price"):
        if row.get(key) not in (None, ""):
            total += float(row[key])
    return total
